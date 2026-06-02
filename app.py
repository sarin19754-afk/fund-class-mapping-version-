"""
app.py  —  Fund Class Mapping  (Streamlit Web App)
รันในเครื่อง:  streamlit run app.py
"""

import io
import sys
from itertools import combinations
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Fund Class Mapping",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────────────────────────────────────
# COLUMN CONFIG DEFAULTS  (แก้ได้จาก sidebar)
# ──────────────────────────────────────────────────────────────────────────────
DEFAULT = dict(
    FA_DATE_COL   = "Effective Date",
    FA_UNIT_COL   = "Alloted Unit",
    FA_CAP_COL    = "CAP",
    FA_CLASS_COL  = "Class",
    SR_DATE_COL   = "DATE",
    SR_UNIT_COL   = "UNITS",
    SR_CAP_COL    = "CAPITAL VALUE",
    TOLERANCE     = 0.01,
    MAX_COMBO_SIZE= 8,
)

# ──────────────────────────────────────────────────────────────────────────────
# CORE FUNCTIONS  (copied from fund_class_mapping.py, I/O adapted for Streamlit)
# ──────────────────────────────────────────────────────────────────────────────

def normalize_dates(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, errors="coerce")
    nat_mask = parsed.isna() & series.notna()
    if nat_mask.any():
        parsed[nat_mask] = pd.to_datetime(series[nat_mask], dayfirst=True, errors="coerce")
    return parsed.dt.normalize()


def to_float(value) -> Optional[float]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def find_subset_indices(values: list, target: float, tol: float, max_size: int) -> Optional[list]:
    n     = len(values)
    limit = min(n, max_size)
    for size in range(1, limit + 1):
        for combo in combinations(range(n), size):
            if abs(sum(values[i] for i in combo) - target) <= tol:
                return list(combo)
    return None


def _build_fa_to_sr_date_map(fa, sr, fa_unit_col, sr_unit_col,
                              search_window_days=10, sum_tol=1.0):
    fa_sums = fa.groupby("_date")[fa_unit_col].sum()
    sr_sums = sr.groupby("_date")[sr_unit_col].sum()
    sr_dates_sorted = sorted(sr_sums.index)
    date_map = {}
    for fa_date, fa_sum in fa_sums.items():
        win_max = fa_date + pd.Timedelta(days=search_window_days)
        for sr_date in sr_dates_sorted:
            if sr_date < fa_date:
                continue
            if sr_date > win_max:
                break
            if abs(sr_sums[sr_date] - fa_sum) <= sum_tol:
                date_map[fa_date] = sr_date
                break
    return date_map


def map_classes(fa_df, sr_df, cfg, progress_cb=None):
    """จับคู่ Class — คืน sr DataFrame พร้อมคอลัมน์ Class และ Match Type"""
    fa = fa_df.copy()
    sr = sr_df.copy()
    sr.columns = sr.columns.str.strip()

    for col in [cfg["FA_UNIT_COL"], cfg["FA_CAP_COL"]]:
        if col in fa.columns:
            fa[col] = fa[col].apply(to_float)
    for col in [cfg["SR_UNIT_COL"], cfg["SR_CAP_COL"]]:
        if col in sr.columns:
            sr[col] = sr[col].apply(to_float)

    fa["_date"] = normalize_dates(fa[cfg["FA_DATE_COL"]])
    sr["_date"] = normalize_dates(sr[cfg["SR_DATE_COL"]])

    sr["Class"]      = "Unmapped"
    sr["Match Type"] = "Unmapped"
    sr["_matched"]   = False

    match_pairs = []
    if cfg["FA_UNIT_COL"] in fa.columns and cfg["SR_UNIT_COL"] in sr.columns:
        match_pairs.append((cfg["FA_UNIT_COL"], cfg["SR_UNIT_COL"], "Units"))
    if cfg["FA_CAP_COL"] in fa.columns and cfg["SR_CAP_COL"] in sr.columns:
        match_pairs.append((cfg["FA_CAP_COL"], cfg["SR_CAP_COL"], "CAP"))

    if not match_pairs:
        raise ValueError("ไม่พบคู่คอลัมน์สำหรับ matching กรุณาตรวจสอบการตั้งค่า")

    fa_to_sr = _build_fa_to_sr_date_map(
        fa, sr, cfg["FA_UNIT_COL"], cfg["SR_UNIT_COL"]
    )

    total = len(fa)
    for i, (_, fa_row) in enumerate(fa.iterrows()):
        if progress_cb:
            progress_cb(i / total)

        fa_date    = fa_row["_date"]
        fund_class = str(fa_row[cfg["FA_CLASS_COL"]]).strip()
        if pd.isna(fa_date):
            continue

        confirmed = fa_to_sr.get(fa_date)
        if confirmed is not None:
            candidate_mask = (sr["_date"] == confirmed) & (~sr["_matched"])
        else:
            from pandas.tseries.offsets import BusinessDay
            sr_min = fa_date + BusinessDay(1)
            sr_max = sr_min + pd.Timedelta(days=3)
            candidate_mask = (
                (sr["_date"] >= sr_min) & (sr["_date"] <= sr_max) & (~sr["_matched"])
            )

        if not candidate_mask.any():
            continue

        for fa_col, sr_col, label in match_pairs:
            target_val = to_float(fa_row.get(fa_col))
            if target_val is None:
                continue
            candidates = sr[candidate_mask & sr[sr_col].notna()].copy()
            if candidates.empty:
                continue
            sr_values   = [to_float(v) for v in candidates[sr_col].values]
            valid_pairs = [(i2, v) for i2, v in enumerate(sr_values) if v is not None]
            if not valid_pairs:
                continue
            valid_positions, valid_vals = zip(*valid_pairs)
            hit = find_subset_indices(list(valid_vals), target_val,
                                      cfg["TOLERANCE"], cfg["MAX_COMBO_SIZE"])
            if hit is not None:
                matched_idx = [candidates.index[valid_positions[p]] for p in hit]
                match_type  = (
                    f"1:1 Match ({label})" if len(hit) == 1
                    else f"Aggregated {len(hit)} orders ({label})"
                )
                sr.loc[matched_idx, "Class"]      = fund_class
                sr.loc[matched_idx, "Match Type"] = match_type
                sr.loc[matched_idx, "_matched"]   = True
                break

    if progress_cb:
        progress_cb(1.0)

    sr.drop(columns=["_date", "_matched"], inplace=True)
    return sr


def build_excel_bytes(df: pd.DataFrame) -> bytes:
    """สร้าง Excel ใน memory แล้วคืนเป็น bytes สำหรับ Streamlit download"""
    import openpyxl
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Mapped Result")
        ws = writer.sheets["Mapped Result"]

        hdr_fill  = PatternFill("solid", fgColor="4472C4")
        hdr_font  = Font(bold=True, color="FFFFFF")
        hdr_align = Alignment(horizontal="center", vertical="center")
        unmap_fill = PatternFill("solid", fgColor="FFD0D0")
        aggr_fill  = PatternFill("solid", fgColor="FFFACD")

        for cell in ws[1]:
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = hdr_align

        col_names = list(df.columns)
        mt_idx    = col_names.index("Match Type") + 1 if "Match Type" in col_names else None
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if mt_idx is None:
                continue
            mv = str(row[mt_idx - 1].value or "")
            fill = unmap_fill if mv == "Unmapped" else (aggr_fill if "Aggregated" in mv else None)
            if fill:
                for cell in row:
                    cell.fill = fill

        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0 for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 60)
        ws.freeze_panes = "A2"

    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ ตั้งค่า")

    st.subheader("คอลัมน์ FA_Trans")
    FA_DATE_COL  = st.text_input("วันที่",         DEFAULT["FA_DATE_COL"])
    FA_UNIT_COL  = st.text_input("จำนวนหน่วย",     DEFAULT["FA_UNIT_COL"])
    FA_CAP_COL   = st.text_input("มูลค่า (CAP)",   DEFAULT["FA_CAP_COL"])
    FA_CLASS_COL = st.text_input("Class",           DEFAULT["FA_CLASS_COL"])

    st.subheader("คอลัมน์ SR")
    SR_DATE_COL  = st.text_input("วันที่",          DEFAULT["SR_DATE_COL"],  key="sr_date")
    SR_UNIT_COL  = st.text_input("จำนวนหน่วย",      DEFAULT["SR_UNIT_COL"],  key="sr_unit")
    SR_CAP_COL   = st.text_input("มูลค่า",          DEFAULT["SR_CAP_COL"],   key="sr_cap")

    st.subheader("พารามิเตอร์การ Match")
    TOLERANCE      = st.number_input("Tolerance (หน่วย)", value=DEFAULT["TOLERANCE"],
                                     min_value=0.0, step=0.001, format="%.4f")
    MAX_COMBO_SIZE = st.slider("Max Combo Size", 1, 12, DEFAULT["MAX_COMBO_SIZE"])

    st.markdown("---")
    st.caption("Fund Class Mapping v1.0")

cfg = dict(
    FA_DATE_COL=FA_DATE_COL, FA_UNIT_COL=FA_UNIT_COL,
    FA_CAP_COL=FA_CAP_COL,   FA_CLASS_COL=FA_CLASS_COL,
    SR_DATE_COL=SR_DATE_COL, SR_UNIT_COL=SR_UNIT_COL,
    SR_CAP_COL=SR_CAP_COL,   TOLERANCE=TOLERANCE,
    MAX_COMBO_SIZE=MAX_COMBO_SIZE,
)

# ──────────────────────────────────────────────────────────────────────────────
# MAIN CONTENT
# ──────────────────────────────────────────────────────────────────────────────
st.title("🏦 Fund Class Mapping")
st.markdown("จับคู่ **Class** กองทุนจาก FA_Trans เข้ากับ Order ใน SR")
st.divider()

# ── Upload section ──
col_up1, col_up2 = st.columns(2)
with col_up1:
    st.subheader("📂 FA_Trans")
    fa_file = st.file_uploader("อัปโหลด FA_Trans.xlsx", type=["xlsx", "xls", "csv"],
                                key="fa_up", label_visibility="collapsed")
    if fa_file:
        st.success(f"✅ {fa_file.name}  ({fa_file.size/1024:.1f} KB)")

with col_up2:
    st.subheader("📂 SR")
    sr_file = st.file_uploader("อัปโหลด SR.xlsx", type=["xlsx", "xls", "csv"],
                                key="sr_up", label_visibility="collapsed")
    if sr_file:
        st.success(f"✅ {sr_file.name}  ({sr_file.size/1024:.1f} KB)")

st.divider()

# ── Preview uploaded files ──
if fa_file or sr_file:
    with st.expander("🔍 ตัวอย่างข้อมูล Input (5 แถวแรก)", expanded=False):
        if fa_file:
            try:
                fa_prev = pd.read_excel(fa_file, nrows=5) if fa_file.name.endswith(("xlsx","xls")) \
                          else pd.read_csv(fa_file, nrows=5)
                fa_file.seek(0)
                st.markdown("**FA_Trans**")
                st.dataframe(fa_prev, use_container_width=True)
            except Exception as e:
                st.error(f"อ่าน FA_Trans ไม่ได้: {e}")
        if sr_file:
            try:
                sr_prev = pd.read_excel(sr_file, nrows=5) if sr_file.name.endswith(("xlsx","xls")) \
                          else pd.read_csv(sr_file, nrows=5)
                sr_file.seek(0)
                st.markdown("**SR**")
                st.dataframe(sr_prev, use_container_width=True)
            except Exception as e:
                st.error(f"อ่าน SR ไม่ได้: {e}")

# ── Run button ──
run_ready = fa_file is not None and sr_file is not None
if st.button("▶ Run Mapping", type="primary", disabled=not run_ready,
             use_container_width=True):

    with st.spinner("กำลังโหลดไฟล์..."):
        try:
            fa_file.seek(0); sr_file.seek(0)
            fa_df = pd.read_excel(fa_file, dtype=str) if fa_file.name.endswith(("xlsx","xls")) \
                    else pd.read_csv(fa_file, dtype=str)
            sr_df = pd.read_excel(sr_file, dtype=str) if sr_file.name.endswith(("xlsx","xls")) \
                    else pd.read_csv(sr_file, dtype=str)
        except Exception as e:
            st.error(f"❌ โหลดไฟล์ไม่สำเร็จ: {e}")
            st.stop()

    # Validate columns
    errors = []
    for col in [cfg["FA_DATE_COL"], cfg["FA_UNIT_COL"], cfg["FA_CLASS_COL"]]:
        if col not in fa_df.columns:
            errors.append(f"FA_Trans: ไม่พบคอลัมน์ '{col}'")
    for col in [cfg["SR_DATE_COL"], cfg["SR_UNIT_COL"]]:
        if col not in sr_df.columns and col not in sr_df.columns.str.strip().tolist():
            errors.append(f"SR: ไม่พบคอลัมน์ '{col}'")
    if errors:
        for e in errors:
            st.error(f"❌ {e}")
        st.info(f"คอลัมน์ใน FA_Trans: {list(fa_df.columns)}")
        st.info(f"คอลัมน์ใน SR: {list(sr_df.columns)}")
        st.stop()

    # Run mapping with progress bar
    prog_bar = st.progress(0, text="กำลัง Map Class...")

    def update_progress(v):
        prog_bar.progress(min(v, 1.0), text=f"กำลัง Map Class... {v*100:.0f}%")

    try:
        result_df = map_classes(fa_df, sr_df, cfg, progress_cb=update_progress)
    except Exception as e:
        st.error(f"❌ เกิดข้อผิดพลาด: {e}")
        st.stop()

    prog_bar.progress(1.0, text="เสร็จสิ้น!")

    # ── Summary metrics ──
    st.divider()
    st.subheader("📊 สรุปผล")
    total    = len(result_df)
    mapped   = (result_df["Class"] != "Unmapped").sum()
    unmapped = total - mapped
    one2one  = result_df["Match Type"].str.startswith("1:1").sum()    if "Match Type" in result_df.columns else 0
    aggr     = result_df["Match Type"].str.startswith("Aggregated").sum() if "Match Type" in result_df.columns else 0

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Orders ทั้งหมด",    f"{total:,}")
    m2.metric("Match สำเร็จ",      f"{mapped:,}", f"{mapped/total*100:.1f}%")
    m3.metric("Unmapped",          f"{unmapped:,}",
              delta=None if unmapped == 0 else f"-{unmapped/total*100:.1f}%",
              delta_color="inverse")
    m4.metric("1:1 Match",         f"{one2one:,}")
    m5.metric("Aggregated Match",  f"{aggr:,}")

    # Class distribution chart
    if mapped > 0:
        st.markdown("**Class Distribution**")
        class_counts = (result_df[result_df["Class"] != "Unmapped"]["Class"]
                        .value_counts().reset_index())
        class_counts.columns = ["Class", "Orders"]
        col_chart, col_tbl = st.columns([2, 1])
        col_chart.bar_chart(class_counts.set_index("Class"), use_container_width=True)
        col_tbl.dataframe(class_counts, use_container_width=True, hide_index=True)

    # ── Result table ──
    st.divider()
    st.subheader("📋 ผลลัพธ์")

    filter_opts = ["ทั้งหมด", "Match สำเร็จ", "Unmapped"]
    filter_sel  = st.radio("แสดง:", filter_opts, horizontal=True)

    if filter_sel == "Match สำเร็จ":
        show_df = result_df[result_df["Class"] != "Unmapped"]
    elif filter_sel == "Unmapped":
        show_df = result_df[result_df["Class"] == "Unmapped"]
    else:
        show_df = result_df

    st.dataframe(
        show_df.head(500),
        use_container_width=True,
        height=400,
        hide_index=True,
    )
    if len(show_df) > 500:
        st.caption(f"แสดง 500 จาก {len(show_df):,} แถว — ดาวน์โหลด Excel เพื่อดูทั้งหมด")

    # ── Download button ──
    st.divider()
    with st.spinner("กำลังสร้างไฟล์ Excel..."):
        excel_bytes = build_excel_bytes(result_df)

    st.download_button(
        label="⬇️ ดาวน์โหลด output_mapped.xlsx",
        data=excel_bytes,
        file_name="output_mapped.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

elif not run_ready:
    st.info("📌 กรุณาอัปโหลดไฟล์ FA_Trans และ SR ทั้งคู่ก่อนกด Run")
